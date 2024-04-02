import openpyxl
import sys

sys.path.append("/Users/sudhansu/Desktop/Folder/My Projects/Selenium project/seleniumPytest/TestData")


class TestData:

    def __init__(self):

        self.file = openpyxl.load_workbook(
            '/Users/sudhansu/Desktop/Folder/My Projects/Selenium project/seleniumPytest/TestData/Add user.xlsx')
        self.sheet = self.file['Sheet1']
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column
        print(f'rows:{self.rows},col:{self.cols}')

    def load_excel_file(self, r, c,sheet):
        dict = {}
        print(f'The row is {r}')
        for j in range(1, c + 1):
            if j == 1:
                dict['User Role'] = sheet.cell(row=r, column=j).value
            elif j == 2:
                dict['Employee Name'] = sheet.cell(row=r, column=j).value
            elif j == 3:
                dict['Status'] = sheet.cell(row=r, column=j).value
            elif j == 4:
                dict['Username'] = sheet.cell(row=r, column=j).value
            elif j == 5:
                dict['Password'] = sheet.cell(row=r, column=j).value
            elif j == 6:
                dict['Confirm Password'] = sheet.cell(row=r, column=j).value
        return dict
